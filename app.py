import threading
import time
import telebot
import sqlite3
import logging
from datetime import datetime, timedelta
import csv
from io import StringIO
import os
from flask import Flask
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ========== НАСТРОЙКИ ==========
TOKEN = "8540729323:AAGzJc9Lv8_Vvd-n7wghRKt10UjV8QL68U0"
ADMIN_ID = "353053358"

# ========== ИНИЦИАЛИЗАЦИЯ ==========
bot = telebot.TeleBot(TOKEN)
app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ========== БАЗА ДАННЫХ ==========
def init_db():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS records
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  date TEXT,
                  object TEXT,
                  employee TEXT,
                  task TEXT,
                  created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS employees (name TEXT PRIMARY KEY)''')
    c.execute('''CREATE TABLE IF NOT EXISTS objects (name TEXT PRIMARY KEY)''')
    for emp in ['Пивненко', 'Дорохин', 'Кравцов']:
        c.execute('INSERT OR IGNORE INTO employees (name) VALUES (?)', (emp,))
    for obj in ['Малайзия', 'Ростех', 'Офис', 'Чистые пруды']:
        c.execute('INSERT OR IGNORE INTO objects (name) VALUES (?)', (obj,))
    conn.commit()
    conn.close()

init_db()

# ========== ФУНКЦИИ БАЗЫ ДАННЫХ ==========
def get_employees():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT name FROM employees ORDER BY name')
    result = [row[0] for row in c.fetchall()]
    conn.close()
    return result

def get_objects():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT name FROM objects ORDER BY name')
    result = [row[0] for row in c.fetchall()]
    conn.close()
    return result

def add_employee(name):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('INSERT OR IGNORE INTO employees (name) VALUES (?)', (name,))
    conn.commit()
    conn.close()

def add_object(name):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('INSERT OR IGNORE INTO objects (name) VALUES (?)', (name,))
    conn.commit()
    conn.close()

def delete_employee(name):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM employees WHERE name = ?', (name,))
    conn.commit()
    conn.close()

def delete_object(name):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM objects WHERE name = ?', (name,))
    conn.commit()
    conn.close()

def add_record(employee, object_name, date_str, task):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('''INSERT INTO records (date, object, employee, task, created_at)
                 VALUES (?, ?, ?, ?, ?)''',
              (date_str, object_name, employee, task, created_at))
    conn.commit()
    conn.close()
    return f"✅ Сохранено: {employee} → {object_name}, {date_str} — {task}"

def get_records_by_date(date_str):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT date, object, employee, task FROM records WHERE date = ? ORDER BY created_at', (date_str,))
    result = [{'date': row[0], 'object': row[1], 'employee': row[2], 'task': row[3] or '—'} for row in c.fetchall()]
    conn.close()
    return result

def get_records_by_period(start_date, end_date):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT date, object, employee, task, created_at FROM records')
    rows = c.fetchall()
    conn.close()
    start = datetime.strptime(start_date, '%d.%m')
    end = datetime.strptime(end_date, '%d.%m')
    result = []
    for row in rows:
        try:
            record_date = datetime.strptime(row[0], '%d.%m')
            if start <= record_date <= end:
                result.append({
                    'date': row[0],
                    'object': row[1],
                    'employee': row[2],
                    'task': row[3] or '—',
                    'created_at': row[4]
                })
        except:
            try:
                record_date = datetime.strptime(row[0], '%d.%m.%Y')
                if start <= record_date <= end:
                    result.append({
                        'date': row[0],
                        'object': row[1],
                        'employee': row[2],
                        'task': row[3] or '—',
                        'created_at': row[4]
                    })
            except:
                pass
    result.sort(key=lambda x: (datetime.strptime(x['date'], '%d.%m'), x.get('created_at', '')))
    return result

def get_stats_by_period(start_date, end_date):
    records = get_records_by_period(start_date, end_date)
    objects = get_objects()
    employees_in_period = list(set(r['employee'] for r in records))
    stats = {}
    for emp in employees_in_period:
        stats[emp] = {obj: 0 for obj in objects}
    for record in records:
        emp = record['employee']
        obj = record['object']
        if emp in stats and obj in stats[emp]:
            stats[emp][obj] += 1
    return stats

def get_all_records():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT date, object, employee, task FROM records ORDER BY date, created_at')
    result = [{'date': row[0], 'object': row[1], 'employee': row[2], 'task': row[3] or '—'} for row in c.fetchall()]
    conn.close()
    return result

def get_all_records_with_ids():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT id, date, object, employee, task FROM records ORDER BY date, created_at')
    result = [{'id': row[0], 'date': row[1], 'object': row[2], 'employee': row[3], 'task': row[4] or '—'} for row in c.fetchall()]
    conn.close()
    return result

def delete_last_record():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM records WHERE id = (SELECT MAX(id) FROM records)')
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted

def delete_record_by_id(record_id):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM records WHERE id = ?', (record_id,))
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted

def filter_records_by_employees(records, selected_employees):
    if not selected_employees or 'Все' in selected_employees:
        return records
    return [r for r in records if r['employee'] in selected_employees]

# ========== ЭКСПОРТ В EXCEL ==========
def generate_detailed_excel(records, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Детальный"
    headers = ['Дата', 'Объект', 'Сотрудник', 'Что делали']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for record in records:
        ws.append([record['date'], record['object'], record['employee'], record['task']])
    for row in ws.iter_rows(min_row=1, max_row=len(records) + 1, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 22
    filename = f"Детальный_{start_date}-{end_date}.xlsx"
    wb.save(filename)
    return filename

def generate_summary_excel(records, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    stats = get_stats_by_period(start_date, end_date)
    objects = get_objects()
    headers = ['Сотрудник'] + objects
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for emp, emp_stats in stats.items():
        row = [emp] + [emp_stats.get(obj, 0) for obj in objects]
        ws.append(row)
    for row in ws.iter_rows(min_row=1, max_row=len(stats) + 1, min_col=1, max_col=len(objects) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    ws.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 14
    filename = f"Сводка_{start_date}-{end_date}.xlsx"
    wb.save(filename)
    return filename

# ========== ГЛАВНОЕ МЕНЮ ==========
def get_main_menu():
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('📋 Добавить запись', '📊 Отчёт')
    markup.add('📅 Весь период', '📅 Неделя')
    markup.add('📅 Месяц', '📅 Свой период')
    markup.add('👤 Сотрудники', '🏗 Объекты')
    markup.add('📋 Повторить сегодня', '↩️ Отменить')
    markup.add('❌ Удалить запись')
    return markup

# ========== ОБРАБОТЧИКИ ==========
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    bot.reply_to(message,
                 "👋 Привет! Я бот для учёта рабочего времени.\n\nНажмите '📋 Добавить запись', чтобы отметить день.",
                 reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '📋 Добавить запись')
def start_add_record(message):
    employees = get_employees()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    for emp in employees:
        markup.add(emp)
    markup.add('➕ Новый сотрудник')
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, "Выберите сотрудника:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_employee)

def process_employee(message):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    employee = message.text
    if employee == '➕ Новый сотрудник':
        msg = bot.reply_to(message, "Введите имя нового сотрудника:")
        bot.register_next_step_handler(msg, process_new_employee)
        return
    if employee not in get_employees():
        bot.reply_to(message, "❌ Сотрудник не найден.")
        return
    objects = get_objects()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    for obj in objects:
        markup.add(obj)
    markup.add('➕ Новый объект')
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, f"Выбран: {employee}\nВыберите объект:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_object, employee)

def process_new_employee(message):
    employee = message.text.strip()
    add_employee(employee)
    bot.reply_to(message, f"✅ {employee} добавлен!")
    objects = get_objects()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    for obj in objects:
        markup.add(obj)
    markup.add('➕ Новый объект')
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, f"Выбран: {employee}\nВыберите объект:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_object, employee)

def process_object(message, employee):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    object_name = message.text
    if object_name == '➕ Новый объект':
        msg = bot.reply_to(message, "Введите название нового объекта:")
        bot.register_next_step_handler(msg, process_new_object, employee)
        return
    if object_name not in get_objects():
        bot.reply_to(message, "❌ Объект не найден.")
        return
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('📅 Сегодня')
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, f"{employee} → {object_name}\nВведите дату (ДД.ММ):", reply_markup=markup)
    bot.register_next_step_handler(msg, process_date, employee, object_name)

def process_new_object(message, employee):
    object_name = message.text.strip()
    add_object(object_name)
    bot.reply_to(message, f"✅ {object_name} добавлен!")
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('📅 Сегодня')
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, f"{employee} → {object_name}\nВведите дату (ДД.ММ):", reply_markup=markup)
    bot.register_next_step_handler(msg, process_date, employee, object_name)

def process_date(message, employee, object_name):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    if message.text == '📅 Сегодня':
        date_str = datetime.now().strftime('%d.%m')
    else:
        date_str = message.text.strip()
        try:
            datetime.strptime(date_str, '%d.%m')
        except:
            bot.reply_to(message, "❌ Неверный формат. Используйте ДД.ММ")
            return
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT * FROM records WHERE date = ? AND employee = ? AND object = ?',
              (date_str, employee, object_name))
    existing = c.fetchone()
    conn.close()
    if existing:
        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('🔄 Заменить', '⏩ Пропустить', '❌ Отмена')
        msg = bot.reply_to(message,
                           f"⚠️ Запись на {date_str} для {employee} на объекте {object_name} уже существует.\nЧто делаем?",
                           reply_markup=markup)
        bot.register_next_step_handler(msg, process_duplicate, employee, object_name, date_str)
        return
    msg = bot.reply_to(message, "Что делали? (кратко):")
    bot.register_next_step_handler(msg, process_task, employee, object_name, date_str)

def process_duplicate(message, employee, object_name, date_str):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    if message.text == '🔄 Заменить':
        conn = sqlite3.connect('data.db')
        c = conn.cursor()
        c.execute('DELETE FROM records WHERE date = ? AND employee = ? AND object = ?',
                  (date_str, employee, object_name))
        conn.commit()
        conn.close()
        msg = bot.reply_to(message, "Введите, что делали:")
        bot.register_next_step_handler(msg, process_task, employee, object_name, date_str)
    elif message.text == '⏩ Пропустить':
        bot.reply_to(message, "⏩ Пропущено.", reply_markup=get_main_menu())
    else:
        bot.reply_to(message, "❌ Неверный выбор.")

def process_task(message, employee, object_name, date_str):
    task = message.text.strip() or '—'
    result = add_record(employee, object_name, date_str, task)
    bot.reply_to(message, result, reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '📊 Отчёт')
def report_custom(message):
    msg = bot.reply_to(message, "Введите период (ДД.ММ-ДД.ММ):")
    bot.register_next_step_handler(msg, process_report_period)

def process_report_period(message, start_date=None, end_date=None):
    try:
        if start_date is None or end_date is None:
            parts = message.text.strip().split('-')
            start_date = parts[0].strip()
            end_date = parts[1].strip()
        records = get_records_by_period(start_date, end_date)
        if not records:
            bot.reply_to(message, "❌ За этот период записей нет.", reply_markup=get_main_menu())
            return
        ask_employees_for_report(message, start_date, end_date)
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {e}\nИспользуйте ДД.ММ-ДД.ММ", reply_markup=get_main_menu())

def ask_employees_for_report(message, start_date, end_date):
    employees = get_employees()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('👥 Все')
    for emp in employees:
        markup.add(emp)
    markup.add('❌ Отмена')
    msg = bot.reply_to(
        message,
        f"👥 Выберите сотрудников для отчёта за {start_date}-{end_date}:\n(можно выбрать несколько, нажимая по очереди)\n\nКогда закончите — нажмите '✅ Готово'",
        reply_markup=markup
    )
    bot.register_next_step_handler(msg, process_employee_selection, start_date, end_date, [])

def process_employee_selection(message, start_date, end_date, selected_employees):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отчёт отменён.", reply_markup=get_main_menu())
        return
    if message.text == '👥 Все':
        selected_employees = ['Все']
        bot.reply_to(message, f"✅ Выбраны все сотрудники. Формирую отчёт...", reply_markup=get_main_menu())
        generate_and_send_report(message, start_date, end_date, selected_employees)
        return
    if message.text == '✅ Готово':
        if not selected_employees:
            bot.reply_to(message, "❌ Вы не выбрали ни одного сотрудника. Попробуйте снова.",
                         reply_markup=get_main_menu())
            return
        bot.reply_to(message, f"✅ Выбраны: {', '.join(selected_employees)}. Формирую отчёт...",
                     reply_markup=get_main_menu())
        generate_and_send_report(message, start_date, end_date, selected_employees)
        return
    if message.text in get_employees():
        if message.text not in selected_employees:
            selected_employees.append(message.text)
            bot.reply_to(message, f"✅ Добавлен: {message.text}\nВыбраны: {', '.join(selected_employees)}")
        else:
            bot.reply_to(message, f"ℹ️ {message.text} уже выбран")
        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('👥 Все')
        for emp in get_employees():
            markup.add(emp)
        markup.add('✅ Готово')
        markup.add('❌ Отмена')
        msg = bot.reply_to(
            message,
            f"Выбраны: {', '.join(selected_employees)}\n\nДобавьте ещё или нажмите '✅ Готово'",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_employee_selection, start_date, end_date, selected_employees)
    else:
        bot.reply_to(message, "❌ Неверный выбор. Попробуйте снова.")

def generate_and_send_report(message, start_date, end_date, selected_employees):
    try:
        all_records = get_records_by_period(start_date, end_date)
        if not all_records:
            bot.reply_to(message, "❌ За этот период записей нет.", reply_markup=get_main_menu())
            return
        if 'Все' not in selected_employees:
            records = filter_records_by_employees(all_records, selected_employees)
        else:
            records = all_records
        if not records:
            bot.reply_to(message, f"❌ Нет записей для выбранных сотрудников.", reply_markup=get_main_menu())
            return
        detailed_path = generate_detailed_excel(records, start_date, end_date)
        summary_path = generate_summary_excel(records, start_date, end_date)
        with open(detailed_path, 'rb') as f:
            bot.send_document(message.chat.id, f, caption=f"📋 Детальный отчёт за {start_date}-{end_date}")
        with open(summary_path, 'rb') as f:
            bot.send_document(message.chat.id, f, caption=f"📊 Сводка за {start_date}-{end_date}")
        os.remove(detailed_path)
        os.remove(summary_path)
        bot.reply_to(message, "✅ Готово! Файлы выше.", reply_markup=get_main_menu())
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {e}", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '📅 Весь период')
def report_all(message):
    records = get_all_records()
    if not records:
        bot.reply_to(message, "❌ Записей нет.", reply_markup=get_main_menu())
        return
    dates = sorted(set(r['date'] for r in records))
    start_date = dates[0]
    end_date = dates[-1]
    process_report_period(message, start_date, end_date)

@bot.message_handler(func=lambda message: message.text == '📅 Неделя')
def report_week(message):
    today = datetime.now()
    week_ago = today - timedelta(days=7)
    start_date = week_ago.strftime('%d.%m')
    end_date = today.strftime('%d.%m')
    process_report_period(message, start_date, end_date)

@bot.message_handler(func=lambda message: message.text == '📅 Месяц')
def report_month(message):
    today = datetime.now()
    month_ago = today - timedelta(days=30)
    start_date = month_ago.strftime('%d.%m')
    end_date = today.strftime('%d.%m')
    process_report_period(message, start_date, end_date)

@bot.message_handler(func=lambda message: message.text == '👤 Сотрудники')
def employees_menu(message):
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('📋 Список сотрудников', '➕ Добавить сотрудника')
    markup.add('❌ Удалить сотрудника', '🔙 Назад')
    bot.reply_to(message, "👤 Сотрудники:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == '📋 Список сотрудников')
def list_employees(message):
    employees = get_employees()
    bot.reply_to(message, "📋 Сотрудники:\n" + "\n".join(f"• {e}" for e in employees))

@bot.message_handler(func=lambda message: message.text == '➕ Добавить сотрудника')
def add_new_employee(message):
    msg = bot.reply_to(message, "Введите имя:")
    bot.register_next_step_handler(msg, process_new_employee_from_menu)

def process_new_employee_from_menu(message):
    name = message.text.strip()
    add_employee(name)
    bot.reply_to(message, f"✅ {name} добавлен!", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '❌ Удалить сотрудника')
def delete_employee_menu(message):
    employees = get_employees()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    for emp in employees:
        markup.add(emp)
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, "Выберите сотрудника:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_delete_employee)

def process_delete_employee(message):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    delete_employee(message.text)
    bot.reply_to(message, f"✅ {message.text} удалён.", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '🏗 Объекты')
def objects_menu(message):
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('📋 Список объектов', '➕ Добавить объект')
    markup.add('❌ Удалить объект', '🔙 Назад')
    bot.reply_to(message, "🏗 Объекты:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == '📋 Список объектов')
def list_objects(message):
    objects = get_objects()
    bot.reply_to(message, "📋 Объекты:\n" + "\n".join(f"• {o}" for o in objects))

@bot.message_handler(func=lambda message: message.text == '➕ Добавить объект')
def add_new_object(message):
    msg = bot.reply_to(message, "Введите название:")
    bot.register_next_step_handler(msg, process_new_object_from_menu)

def process_new_object_from_menu(message):
    name = message.text.strip()
    add_object(name)
    bot.reply_to(message, f"✅ {name} добавлен!", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '❌ Удалить объект')
def delete_object_menu(message):
    objects = get_objects()
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    for obj in objects:
        markup.add(obj)
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, "Выберите объект:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_delete_object)

def process_delete_object(message):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    delete_object(message.text)
    bot.reply_to(message, f"✅ {message.text} удалён.", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '📋 Повторить сегодня')
def copy_yesterday(message):
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%d.%m')
    records = get_records_by_date(yesterday)
    if not records:
        bot.reply_to(message, "❌ За вчера нет записей.", reply_markup=get_main_menu())
        return
    today = datetime.now().strftime('%d.%m')
    success = []
    for r in records:
        success.append(add_record(r['employee'], r['object'], today, r['task']))
    bot.reply_to(message, "\n".join(success), reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '↩️ Отменить')
def undo_last(message):
    if delete_last_record():
        bot.reply_to(message, "↩️ Последняя запись удалена.", reply_markup=get_main_menu())
    else:
        bot.reply_to(message, "❌ Нет записей.", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '❌ Удалить запись')
def delete_record_menu(message):
    records = get_all_records_with_ids()
    if not records:
        bot.reply_to(message, "❌ Записей нет.", reply_markup=get_main_menu())
        return
    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    for r in records[:10]:
        markup.add(f"🗑 {r['date']} {r['employee']} → {r['object']}")
    markup.add('❌ Отмена')
    msg = bot.reply_to(message, "Выберите запись:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_confirm_delete, records)

def process_confirm_delete(message, records):
    if message.text == '❌ Отмена':
        bot.reply_to(message, "❌ Отменено.", reply_markup=get_main_menu())
        return
    selected = message.text.replace('🗑 ', '')
    for r in records:
        if f"{r['date']} {r['employee']} → {r['object']}" == selected:
            if delete_record_by_id(r['id']):
                bot.reply_to(message, "✅ Удалено.", reply_markup=get_main_menu())
                return
    bot.reply_to(message, "❌ Не найдено.", reply_markup=get_main_menu())

@bot.message_handler(func=lambda message: message.text == '🔙 Назад')
def back_to_main(message):
    bot.reply_to(message, "Главное меню:", reply_markup=get_main_menu())
    
@bot.message_handler(commands=['backup'])
def backup_db(message):
    if str(message.chat.id) != ADMIN_ID:
        bot.reply_to(message, "❌ У вас нет доступа к этой команде.")
        return
    try:
        with open('data.db', 'rb') as f:
            bot.send_document(message.chat.id, f, caption="📦 Резервная копия базы данных")
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {e}")

# ========== FLASK ==========
@app.route('/')
def index():
    return 'Бот работает!', 200

# ========== ЗАПУСК БОТА В ПОТОКЕ ==========
def run_bot():
    time.sleep(2)
    bot.remove_webhook()
    print("✅ Бот запущен в отдельном потоке!")
    bot.infinity_polling()

thread = threading.Thread(target=run_bot)
thread.daemon = True
thread.start()

print("✅ Flask и бот запущены!")
